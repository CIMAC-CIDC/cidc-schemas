"""Tests for pipeline config generation."""
from collections import defaultdict
import os
import copy
from tempfile import NamedTemporaryFile
from typing import Dict, List, Tuple, Union
import openpyxl
import pandas as pd
import yaml

import pytest

from cidc_schemas.template import LocalFileUploadEntry, Template
from cidc_schemas.template_reader import XlTemplateReader
from cidc_schemas.prism import constants, core, pipelines, merger
from cidc_schemas.util import participant_id_from_cimac

from ..constants import TEMPLATE_EXAMPLES_DIR
from ..test_templates import (
    # testing fixtures
    template_set,
    template,
    template_example,
    template_example_xlsx_path,
)
from .cidc_test_data import get_test_trial


@pytest.fixture(scope="session")
def prismify_result(template, template_example):

    # tear down
    core._encrypt_hmac = None
    # and set up
    core.set_prism_encrypt_key("test")

    prism_patch, file_maps, errs = core.prismify(template_example, template)
    assert not errs, "\n".join([str(e) for e in errs])
    return prism_patch, file_maps, errs


def prism_patch_stage_artifacts(
    prismify_result: Tuple[
        dict, List[LocalFileUploadEntry], List[Union[Exception, str]]
    ],
    template_type: str,
) -> dict:
    prism_patch: dict
    prism_fmap: List[LocalFileUploadEntry]

    prism_patch, prism_fmap, _ = prismify_result
    patch_copy_4_artifacts: dict = copy.deepcopy(prism_patch)

    patch_copy_4_artifacts, _ = merger.merge_artifacts(
        patch_copy_4_artifacts,
        [
            merger.ArtifactInfo(
                artifact_uuid=fmap_entry.upload_placeholder,
                object_url=fmap_entry.gs_key,
                upload_type=template_type,
                file_size_bytes=i,
                uploaded_timestamp="01/01/2001",
                md5_hash=f"hash_{i}",
            )
            for i, fmap_entry in enumerate(prism_fmap)
        ],
    )

    return patch_copy_4_artifacts


def stage_assay_for_analysis(template_type: str, full_ct: dict) -> dict:
    """
    Simulates an initial assay upload by prismifying the initial assay template object.
    Returns the updated full clinical trial data after merging.
    """

    staging_map: Dict[str, List[str]] = {
        "cytof_analysis": ["cytof"],
        "tumor_normal_pairing": ["wes_bam", "wes_fastq"],
    }

    for to_patch in staging_map.get(template_type, []):
        full_ct = stage_assay(template_type=to_patch, full_ct=full_ct)

    return full_ct


def get_patch_for_staging(template_type: str) -> dict:
    preassay_xlsx_path = os.path.join(
        TEMPLATE_EXAMPLES_DIR, template_type + "_template.xlsx"
    )
    preassay_xlsx, _ = XlTemplateReader.from_excel(preassay_xlsx_path)
    preassay_template = Template.from_type(template_type)
    prism_res: Tuple[
        dict, List[LocalFileUploadEntry], List[Union[Exception, str]]
    ] = core.prismify(preassay_xlsx, preassay_template)

    patch_with_artifacts: dict = prism_patch_stage_artifacts(prism_res, template_type)
    return patch_with_artifacts


def stage_assay(template_type: str, full_ct: dict) -> dict:
    patch_with_artifacts = get_patch_for_staging(template_type)
    if patch_with_artifacts:
        full_ct, errs = merger.merge_clinical_trial_metadata(
            patch_with_artifacts, full_ct
        )
        assert 0 == len(errs), str(errs)

    return full_ct


def test_WES_pipeline_config_generation_after_prismify(prismify_result, template):

    if not (template.type.startswith("wes_") or "pair" in template.type):
        return

    # Test that the config generator blocks disallowed upload types
    upload_type = "foo"
    with pytest.raises(NotImplementedError, match=f"Not supported type:{upload_type}"):
        pipelines._Wes_pipeline_config(upload_type)

    full_ct = get_test_trial(
        [
            "CTTTPP111.00",
            "CTTTPP121.00",
            "CTTTPP122.00",
            "CTTTPP123.00",
            "CTTTPP124.00",
            "CTTTPP211.00",
            "CTTTPP212.00",
            "CTTTPP213.00",
            "CTTTPP214.00",
            "CTTTPP311.00",
            "CTTTPP312.00",
            "CTTTPP313.00",
            "CTTTPP411.00",
            "CTTTPP412.00",
            "CTTTPP413.00",
            "CTTTPP511.00",
            "CTTTPP512.00",
            "CTTTPP513.00",
            "CTTTPP514.00",
            "CTTTPP515.00",
            "CTTTPP516.00",
            "CTTTPP517.00",
            "CTTTPP518.00",
        ],
        allowed_collection_event_names=[
            "Not_reported",
            "Baseline",
            "On_Treatment",
            "Week_1",
        ],
        assays={
            "wes": [],
        },
        analysis={
            "wes_analysis_old": {
                "pair_runs": [
                    # no report so ignored
                    {
                        "run_id": "CTTTPP111.00",
                        "tumor": {"cimac_id": "CTTTPP111.00"},
                        "normal": {"cimac_id": "CTTTPP121.00"},
                    },
                    # triggers CTTTPP123.00 legacy_normal
                    # causes tumor CTTTPP518.00 to be ignored
                    {
                        "run_id": "CTTTPP518.00",
                        "tumor": {"cimac_id": "CTTTPP518.00"},
                        "normal": {"cimac_id": "CTTTPP123.00"},
                        "report": {
                            "report": {"upload_placeholder": "foo"},
                            "config": {"upload_placeholder": "foo"},
                            "metasheet": {"upload_placeholder": "foo"},
                            "report": {"upload_placeholder": "foo"},
                            "wes_run_version": {"upload_placeholder": "foo"},
                            "wes_sample_json": {"upload_placeholder": "foo"},
                        },
                    },
                ]
            },
            # triggers CTTTPP211.00 legacy_tumor_only
            "wes_tumor_only_analysis": {
                "excluded_samples": [
                    {"cimac_id": "CTTTPP517.00", "reason_excluded": "test"}
                ],
                "runs": [
                    {
                        "run_id": "CTTTPP211.00",
                        "tumor": {"cimac_id": "CTTTPP211.00"},
                        "report": {
                            "report": {"upload_placeholder": "foo"},
                            "config": {"upload_placeholder": "foo"},
                            "metasheet": {"upload_placeholder": "foo"},
                            "report": {"upload_placeholder": "foo"},
                            "wes_run_version": {"upload_placeholder": "foo"},
                            "wes_sample_json": {"upload_placeholder": "foo"},
                        },
                    }
                ],
            },
        },
    )
    # manually modify json's to add tumor / normal definitions for WES
    # these are normally loaded from the shipping manifests
    if "wes" in template.type or template.type == "tumor_normal_pairing":
        for partic in full_ct["participants"]:
            partic_id = partic["cimac_participant_id"]
            if partic_id == "CTTTPP1":
                # these are paired in tumor_normal_pairing
                partic["samples"][0]["processed_sample_derivative"] = "Tumor DNA"
                partic["samples"][1]["processed_sample_derivative"] = "Germline DNA"

                # test default to tumor if not specified
                partic["samples"][2]["collection_event_name"] = "Baseline"

                # test deduplication of normals by collection_event_name
                for n in (3, 4):
                    partic["samples"][n]["processed_sample_derivative"] = "Germline DNA"
                    partic["samples"][n]["collection_event_name"] = "Baseline"
            elif partic_id == "CTTTPP2":
                # test 2 tumor samples on treatment and only 1 normal on treatment
                for n in (0, 1):
                    partic["samples"][n]["processed_sample_derivative"] = "Tumor DNA"
                    partic["samples"][n]["collection_event_name"] = "On_Treatment"
                for n in (2, 3):
                    partic["samples"][n]["processed_sample_derivative"] = "Germline DNA"
                partic["samples"][2]["collection_event_name"] = "On_Treatment"
                partic["samples"][3]["collection_event_name"] = "Baseline"
            elif partic_id == "CTTTPP3":
                # test 1 tumor sample not reported and one of normal samples baseline
                partic["samples"][0]["processed_sample_derivative"] = "Tumor DNA"
                for n in (1, 2):
                    partic["samples"][n]["processed_sample_derivative"] = "Germline DNA"
                partic["samples"][1]["collection_event_name"] = "On_Treatment"
                partic["samples"][2]["collection_event_name"] = "Baseline"
            elif partic_id == "CTTTPP4":
                # test 1 tumor sample not reported and 2 normals other collection events
                partic["samples"][0]["processed_sample_derivative"] = "Tumor DNA"
                for n in (1, 2):
                    partic["samples"][n]["processed_sample_derivative"] = "Germline DNA"
                partic["samples"][1]["collection_event_name"] = "On_Treatment"
                partic["samples"][2]["collection_event_name"] = "Week_1"
            elif partic_id == "CTTTPP5":
                # test tumor samples with no paired normal
                for i in range(len(partic["samples"])):
                    partic["samples"][i]["processed_sample_derivative"] = "Tumor DNA"
                    partic["samples"][i]["collection_event_name"] = "On_Treatment"

    # if it's an analysis - we need to merge corresponding preliminary assay first
    full_ct = stage_assay_for_analysis(template.type, full_ct)

    patch_with_artifacts = prism_patch_stage_artifacts(prismify_result, template.type)
    full_ct, errs = merger.merge_clinical_trial_metadata(patch_with_artifacts, full_ct)
    assert 0 == len(errs), str(errs)

    res = pipelines.generate_analysis_configs_from_upload_patch(
        full_ct, patch_with_artifacts, template.type, "my-biofx-bucket"
    )

    if "analysis" in template.type:
        # where we don't expect to have configs
        assert len(res) == 0
        # that's it, nothing in res to check
        return

    elif "wes" in template.type:
        # 1 pairing csv
        assert len(res) == 1

    else:  # if template.type == "tumor_normal_pairing"
        # 2 config, 1 paired plus 20 tumor-only
        # 20 wes_tumor_only_analysis xlsx
        # 1 wes_analysis xlsx
        # 1 pairing csv
        assert len(res) == 24, list(res.keys())

    pairing_filename = full_ct["protocol_identifier"] + "_pairing.csv"
    if template.type == "wes_bam":
        assert (
            res.pop(pairing_filename) == "protocol_identifier,test_prism_trial_id\n"
            "tumor,tumor_collection_event,legacy_tumor_only,previously_excluded,normal,normal_collection_event,legacy_normal\n"
            "CTTTPP111.00,Not_reported,,,CTTTPP121.00,Not_reported,"
        )
    elif template.type == "wes_fastq":
        assert (
            res.pop(pairing_filename) == "protocol_identifier,test_prism_trial_id\n"
            "tumor,tumor_collection_event,legacy_tumor_only,previously_excluded,normal,normal_collection_event,legacy_normal\n"
            "CTTTPP122.00,Baseline,,,CTTTPP123.00,Baseline,TRUE\n"
            "CTTTPP211.00,On_Treatment,TRUE,,CTTTPP213.00,On_Treatment,\n"
            "CTTTPP212.00,On_Treatment,,,CTTTPP213.00,On_Treatment,\n"
            ",,,,CTTTPP214.00,Baseline,\n"
            "CTTTPP311.00,Not_reported,,,CTTTPP313.00,Baseline,\n"
            ",,,,CTTTPP312.00,On_Treatment,\n"
            "CTTTPP411.00,Not_reported,,,,,\n"
            ",,,,CTTTPP412.00,On_Treatment,\n"
            ",,,,CTTTPP413.00,Week_1,\n"
            "CTTTPP511.00,On_Treatment,,,,,\n"
            "CTTTPP512.00,On_Treatment,,,,,\n"
            "CTTTPP513.00,On_Treatment,,,,,\n"
            "CTTTPP514.00,On_Treatment,,,,,\n"
            "CTTTPP515.00,On_Treatment,,,,,\n"
            "CTTTPP516.00,On_Treatment,,,,,\n"
            "CTTTPP517.00,On_Treatment,,TRUE,,,"
        )
    else:  # if template.type == "tumor_normal_pairing"
        assert (
            res.pop(pairing_filename) == "protocol_identifier,test_prism_trial_id\n"
            "tumor,tumor_collection_event,legacy_tumor_only,previously_excluded,normal,normal_collection_event,legacy_normal\n"
            "CTTTPP111.00,Not_reported,,,CTTTPP121.00,Not_reported,\n"
            "CTTTPP122.00,Baseline,,,CTTTPP123.00,Baseline,TRUE\n"
            "CTTTPP211.00,On_Treatment,TRUE,,CTTTPP213.00,On_Treatment,\n"
            "CTTTPP212.00,On_Treatment,,,CTTTPP213.00,On_Treatment,\n"
            ",,,,CTTTPP214.00,Baseline,\n"
            "CTTTPP311.00,Not_reported,,,CTTTPP313.00,Baseline,\n"
            ",,,,CTTTPP312.00,On_Treatment,\n"
            "CTTTPP411.00,Not_reported,,,,,\n"
            ",,,,CTTTPP412.00,On_Treatment,\n"
            ",,,,CTTTPP413.00,Week_1,\n"
            "CTTTPP511.00,On_Treatment,,,,,\n"
            "CTTTPP512.00,On_Treatment,,,,,\n"
            "CTTTPP513.00,On_Treatment,,,,,\n"
            "CTTTPP514.00,On_Treatment,,,,,\n"
            "CTTTPP515.00,On_Treatment,,,,,\n"
            "CTTTPP516.00,On_Treatment,,,,,\n"
            "CTTTPP517.00,On_Treatment,,TRUE,,,"
        )

    # only generated for pairing manifest
    if "pairing" in template.type:
        all_tumor_cimac_ids: List[
            str
        ] = []  # to make sure all are caught across configs
        for fname, conf in res.items():
            # check the ingestion template excels
            if "template" in fname:
                # openpyxl needs to file to have an .xlsx extension to open it
                with NamedTemporaryFile(suffix=".xlsx") as tmp:
                    tmp.write(conf)
                    tmp.seek(0)
                    wb = openpyxl.load_workbook(tmp.name, data_only=True)

                if "WES Analysis" in wb.sheetnames:
                    sht = wb["WES Analysis"]
                elif "WES tumor-only Analysis" in wb.sheetnames:
                    sht = wb["WES tumor-only Analysis"]
                else:
                    assert (
                        False
                    ), f"Attached xlsx doesn't have right worksheets: {wb.sheetnames}"

                assert sht["C2"].value == "test_prism_trial_id"
                assert sht["B7"].value  # run name
                assert sht["C7"].value  # first id

                if sht.title == "WES Analysis":
                    assert sht["D7"].value  # second id
                    assert sht["B7"].value == sht["D7"].value  # run name is tumor id
                else:
                    assert sht["B7"].value == sht["C7"].value  # run name is tumor id

                # loading folder is based on the trial and tumor cimac (ie run) ids
                assert (
                    sht["C3"].value
                    == f"gs://repro_test_prism_trial_id/WES_v3/{sht['B7'].value}/"
                )

            # check the config template excels
            else:  # if "wes_ingestion" in fname
                df = pd.read_excel(conf)
                all_tumor_cimac_ids.extend(df["tumor_cimac_id"].values)
                assert (
                    df["google_bucket_path"]
                    .str.startswith("gs://repro_test_prism_trial_id/WES_v3/")
                    .all()
                )
                assert df["rna_bam_file"].isna().all()
                assert df["rna_expression_file"].isna().all()
                assert (df["cores"] == 64).all()
                assert (df["disk_size"] == 500).all()
                assert (df["wes_commit"] == "21376c4").all()
                assert (df["image"] == "wes-ver3-01c").all()
                assert (df["wes_ref_snapshot"] == "wes-human-ref-ver1-8").all()
                assert (df["somatic_caller"] == "tnscope").all()
                assert (~df["trim_soft_clip"]).all()
                assert (df["zone"] == "us-east1-c").all()

                if df.shape[0] == 20:
                    # first row is paired, so it's special
                    # if we pop it, can use same code below for all tumor-only
                    first_row, df = df.iloc[0], df.iloc[1:]
                    assert not first_row["tumor_only"]
                    assert first_row["tumor_cimac_id"] == "CTTTPP111.00"
                    assert first_row["normal_cimac_id"] == "CTTTPP121.00"
                    assert first_row["tumor_fastq_path_pair1"].startswith(
                        "gs://my-biofx-bucket/test_prism_trial_id/wes/CTTTPP111.00/"
                    ) and first_row["tumor_fastq_path_pair1"].endswith(".bam")
                    assert first_row["normal_fastq_path_pair1"].startswith(
                        "gs://my-biofx-bucket/test_prism_trial_id/wes/CTTTPP121.00/"
                    ) and first_row["normal_fastq_path_pair1"].endswith(".bam")
                    assert first_row["cimac_center"] == "broad"

                    assert pd.isna(first_row["tumor_fastq_path_pair2"]) and pd.isna(
                        first_row["normal_fastq_path_pair2"]
                    )
                else:
                    assert df.shape[0] == 1

                # if popped above, rest of the rows are tumor-only
                # if only one row, should be tumor-only
                assert (
                    df["tumor_fastq_path_pair1"]
                    .str.startswith(
                        "gs://repro_test_prism_trial_id/WES/fastq/concat_all/analysis/concat/"
                    )
                    .all()
                )
                assert (
                    df["tumor_fastq_path_pair2"]
                    .str.startswith(
                        "gs://repro_test_prism_trial_id/WES/fastq/concat_all/analysis/concat/"
                    )
                    .all()
                )
                assert (
                    df[
                        [
                            "normal_cimac_id",
                            "normal_fastq_path_pair1",
                            "normal_fastq_path_pair2",
                        ]
                    ]
                    .isna()
                    .all()
                    .all()
                )
                assert (df["cimac_center"] == "mda").all()
                assert df["tumor_only"].all()

        # outside of `for` ie once per test
        # to make sure all are caught across configs
        assert sorted(all_tumor_cimac_ids) == [
            "CTTTPP111.00",
            "CTTTPP122.00",
            "CTTTPP123.00",
            "CTTTPP124.00",
            "CTTTPP211.00",
            "CTTTPP212.00",
            "CTTTPP213.00",
            "CTTTPP214.00",
            "CTTTPP311.00",
            "CTTTPP312.00",
            "CTTTPP313.00",
            "CTTTPP411.00",
            "CTTTPP412.00",
            "CTTTPP413.00",
            "CTTTPP511.00",
            "CTTTPP512.00",
            "CTTTPP513.00",
            "CTTTPP514.00",
            "CTTTPP515.00",
            "CTTTPP516.00",
            "CTTTPP517.00",
        ]


def test_RNAseq_pipeline_config_generation_after_prismify(prismify_result, template):

    if not template.type.startswith("rna_"):
        return

    full_ct = get_test_trial(
        [
            "CTTTPP111.00",
            "CTTTPP121.00",
            "CTTTPP122.00",
            "CTTTPP123.00",
            "CTTTPP124.00",
            "CTTTPP125.00",
            "CTTTPP126.00",
        ],
        assays={"rna": []},
    )
    patch_with_artifacts = prism_patch_stage_artifacts(prismify_result, template.type)

    # if it's an analysis - we need to merge corresponding preliminary assay first
    full_ct = stage_assay_for_analysis(template.type, full_ct)

    full_ct, errs = merger.merge_clinical_trial_metadata(patch_with_artifacts, full_ct)

    assert 0 == len(errs), "\n".join(errs)

    res = pipelines.generate_analysis_configs_from_upload_patch(
        full_ct, patch_with_artifacts, template.type, "my-biofx-bucket"
    )

    # where we don't expect to have configs
    if not template.type in pipelines._ANALYSIS_CONF_GENERATORS:
        assert res == {}
        return

    if template.type in ["rna_fastq", "rna_bam"]:
        # one config and one ingestion sheet per batch
        # two samples in fastq example .yaml
        # five samples in bam example .yaml
        assert len(res) == 2 * (1 if template.type == "rna_fastq" else 2)

    else:
        assert False, f"Unexpected RNAseq template test {template.type}"

    # separate batches by int for comparison
    batches: Dict[int, Dict[str, Union[bytes, str]]] = defaultdict(dict)
    for fname, fcontent in res.items():
        # one of:
        # f"rna_ingestion_{trial_id}.batch_{batch_num}_of_{num_batches}.{timestamp}.xlsx"
        # f"rna_pipeline_{trial_id}.batch_{batch_num}_of_{num_batches}.{timestamp}.yaml"
        batch_num: int = int(fname.split(".")[1].split("_")[1])
        batches[batch_num][fname] = fcontent

    total_samples: int = 0
    for batch_num, batch in batches.items():
        assert len(batch) == 2  # yaml, xlsx
        yaml_keys: List[str] = [
            fname for fname in batch.keys() if fname.endswith(".yaml")
        ]
        excel_keys: List[str] = [
            fname for fname in batch.keys() if fname.endswith(".xlsx")
        ]
        assert len(yaml_keys) == 1 and len(excel_keys) == 1

        yaml_content: str = batch[yaml_keys[0]]
        excel_content: bytes = batch[excel_keys[0]]

        # --- Check the ingestion sheet --- #
        # openpyxl needs to file to have an .xlsx extension to open it
        with NamedTemporaryFile(suffix=".xlsx") as tmp_excel:
            tmp_excel.write(excel_content)
            tmp_excel.seek(0)
            wb = openpyxl.load_workbook(tmp_excel.name, data_only=True)

        assert sum([ws.lower().startswith("rna") for ws in wb.sheetnames]) == 1
        sht = wb[[ws for ws in wb.sheetnames if ws.lower().startswith("rna")][0]]
        assert sht["C2"].value == "test_prism_trial_id"
        assert sht["C3"].value == pipelines.RNA_INGESTION_FOLDER

        if template.type == "rna_fastq":
            cimac_ids = [sht["B7"].value, sht["B8"].value]
            assert sht["B9"].value in ["", None]
            assert sorted(cimac_ids) == ["CTTTPP122.00", "CTTTPP123.00"]
        else:  # if template.type == "rna_bam"
            cimac_ids = [
                c
                for c in [
                    sht["B7"].value,
                    sht["B8"].value,
                    sht["B9"].value,
                    sht["B10"].value,
                ]
                if c
            ]
            assert sht["B11"].value in ["", None]
            assert sorted(cimac_ids) in (
                ["CTTTPP122.00", "CTTTPP123.00", "CTTTPP124.00", "CTTTPP125.00"],
                ["CTTTPP126.00"],
            )

        # --- Check the configuration sheet --- #
        conf = yaml.load(yaml_content, Loader=yaml.FullLoader)
        for key in [
            "instance_name",
            "cores",
            "disk_size",
            "google_bucket_path",
            "rima_commit",
            "serviceAcct",
            "image",
            "rima_ref_snapshot",
        ]:
            assert key in conf, key

        total_samples += len(conf["samples"])
        for sample in conf["samples"].values():
            # at least one bam or two fastq file per sample
            assert len(sample) > (1 if template.type == "rna_fastq" else 0)
            assert all("my-biofx-bucket" in f for f in sample)
            assert all(f.endswith(".fastq.gz") for f in sample) or all(
                f.endswith(".bam") for f in sample
            )

        for cimac_id, sample in conf["metasheet"].items():
            assert sample["SampleName"] == cimac_id
            assert sample["PatName"] == participant_id_from_cimac(cimac_id)

        if template.type == "rna_fastq":
            assert len(conf["samples"]) == 2  # two samples in example .xlsx
        else:  # if template.type == "rna_bam"
            assert len(conf["samples"]) in (1, 4)  # five samples in example .xlsx

    # outside of `for`, so after all is done
    if template.type == "rna_fastq":
        assert total_samples == 2  # two samples in example .xlsx
    else:  # if template.type == "rna_bam"
        assert total_samples == 5  # five samples in example .xlsx


def test_TCRseq_pipeline_config_generation_after_prismify(prismify_result, template):

    if not template.type.startswith("tcr_"):
        return

    full_ct = get_test_trial(
        [
            "CTTTPP111.00",
            "CTTTPP121.00",
            "CTTTPP122.00",
        ],
        assays={"tcr": []},
    )
    patch_with_artifacts = prism_patch_stage_artifacts(prismify_result, template.type)

    # if it's an analysis - we need to merge corresponding preliminary assay first
    full_ct = stage_assay_for_analysis(template.type, full_ct)

    full_ct, errs = merger.merge_clinical_trial_metadata(patch_with_artifacts, full_ct)

    print(prismify_result)
    assert 0 == len(errs), "\n".join(errs)

    res = pipelines.generate_analysis_configs_from_upload_patch(
        full_ct, patch_with_artifacts, template.type, "my-biofx-bucket"
    )

    # where we don't expect to have configs
    if not template.type in pipelines._ANALYSIS_CONF_GENERATORS:
        assert res == {}
        return

    # one meta csv batch
    assert len(res) == 1

    for fname, fcontent in res.items():
        assert fname.endswith(".csv")

        assert fcontent == "sample,batch" "\nCTTTPP111.00,XYZ" "\nCTTTPP121.00,XYZ"


def test_shipping_manifest_new_participants_after_prismify(prismify_result, template):

    if not template.type in constants.SUPPORTED_SHIPPING_MANIFESTS:
        return

    base_ct = get_test_trial(
        allowed_cohort_names=["Arm_A", "Arm_Z"],
        allowed_collection_event_names=[
            "Baseline",
            "Pre_Day_1_Cycle_2",
        ],
    )

    patch_with_artifacts: dict = prism_patch_stage_artifacts(
        prismify_result, template.type
    )
    full_ct, errs = merger.merge_clinical_trial_metadata(patch_with_artifacts, base_ct)
    assert 0 == len(errs), "\n".join(errs)

    # test returns all participants on first upload
    res = pipelines.generate_analysis_configs_from_upload_patch(
        full_ct, patch_with_artifacts, template.type, "my-biofx-bucket"
    )
    assert len(res) == 1

    expected_answer = {
        "h_and_e": ["CTTTP08"],
        "microbiome_dna": ["CTTTP08", "CTTTP09"],
        "normal_blood_dna": ["CTTTP01", "CTTTP02"],
        "normal_tissue_dna": ["CTTTP03", "CTTTP04"],
        "pbmc": ["CTTTP01", "CTTTP02"],
        "plasma": ["CTTTP01", "CTTTP02"],
        "tissue_slide": ["CTTTP08", "CTTTP09"],
        "tumor_tissue_dna": ["CTTTP05", "CTTTP06"],
        "tumor_tissue_rna": ["CTTTP05", "CTTTP06"],
    }
    assert res["new_participants.txt"].split("\n") == expected_answer[template.type]

    # test ONLY new participants on subset
    if template.type == "h_and_e":
        patch_with_artifacts = get_patch_for_staging(template_type="microbiome_dna")
        full_ct, errs = merger.merge_clinical_trial_metadata(
            patch_with_artifacts, full_ct
        )
        assert 0 == len(errs), "\n".join(errs)

        res = pipelines.generate_analysis_configs_from_upload_patch(
            full_ct, patch_with_artifacts, template.type, "my-biofx-bucket"
        )
        assert len(res) == 1

        # doesn't return the one from h_and_e
        assert (
            res["new_participants.txt"].split("\n")
            == expected_answer["microbiome_dna"][1:]
        )

    # test doesn't return if no new particpants
    if template.type == "pbmc":
        patch_with_artifacts = get_patch_for_staging(template_type="plasma")
        full_ct, errs = merger.merge_clinical_trial_metadata(
            patch_with_artifacts, full_ct
        )
        assert 0 == len(errs), "\n".join(errs)

        res = pipelines.generate_analysis_configs_from_upload_patch(
            full_ct, patch_with_artifacts, template.type, "my-biofx-bucket"
        )
        assert res == dict()
