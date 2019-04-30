#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""Tests for `cidc_prism` package."""

import os
import unittest
import pytest
import jsonschema
from openpyxl import load_workbook

from cidc_schemas import prism

CUR_DIR = os.path.dirname(os.path.realpath(__file__))

class Test_prism(unittest.TestCase):
  """Tests for `cidc_prism` package."""

  def test_datefields(self):
    """ test parsing support for different datetimes """

    # create schema
    schema = {
      'id': 'schema1',
      'title': 'Schema1',
      'type': 'object',
      'description': 'Some text.',
      'properties': {
        'slashes': {
          'description': 'text',
          'type': 'string',
          'format': 'date'
        },
        'slashes-txt': {
          'description': 'text',
          'type': 'string',
          'format': 'date'
        }
      }
    }

    # build coercion functions
    coercion = {}
    for key in schema['properties']:
      prism.determine_coercion(schema, key, coercion)

    # load the demo workbook and test examples.
    file_path = os.path.join(CUR_DIR, 'data/date_examples.xlsx')
    wb = load_workbook(file_path)
    names = wb.sheetnames
    sheet = wb[names[0]]

    # loop over rows
    instance = {}
    for row in sheet.iter_rows():
      
      # get data.
      line = []
      for col in row:
        line.append(col.value)

      key = line[0]
      val = line[1]

      # assert test is valid
      assert key in schema['properties']

      # save it
      instance[key] = coercion[key](val)

    # validate instance
    prism.validate_schema(schema, instance)

    
  def test_pbcm(self):
    """test static pbmc"""

    # test pbmc
    path_to_manifest = os.path.join(CUR_DIR, 'data/pbmc_shipping.xlsx')
    head_objs, data_objs = prism.validate_instance(path_to_manifest)

    assert len(head_objs) > 0, 'header data is missing'
    assert len(data_objs) > 0, 'row data is missing'