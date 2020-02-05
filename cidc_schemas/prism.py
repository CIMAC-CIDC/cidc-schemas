import re
import json
import os
import copy
import uuid
import datetime
from typing import (
    Union,
    BinaryIO,
    Tuple,
    List,
    NamedTuple,
    Any,
    Optional,
    Dict,
    Callable,
)

import openpyxl
import jsonschema
import jinja2
from jsonmerge import merge, Merger, exceptions as jsonmerge_exceptions, strategies
from collections import namedtuple
from jsonpointer import JsonPointer, JsonPointerException, resolve_pointer, EndOfList

from cidc_schemas.json_validation import load_and_validate_schema
from cidc_schemas.template import Template
from cidc_schemas.template_writer import RowType
from cidc_schemas.template_reader import XlTemplateReader

from cidc_schemas.constants import SCHEMA_DIR, TEMPLATE_DIR
from .util import get_path, get_source

PROTOCOL_ID_FIELD_NAME = "protocol_identifier"


def _set_val(
    pointer: str,
    val: object,
    context: dict,
    root: Union[dict, None] = None,
    context_pointer: Union[str, None] = None,
    verb=False,
):
    """
    This function given a *pointer* (jsonpointer RFC 6901 or relative json pointer)
    to a property in a python object, sets the supplied value
    in-place in the *context* object within *root* object.
    The object we are adding data to is *root*. The object
    may or may not have any of the intermediate structure
    to fully insert the desired property.
    For example: consider
    pointer = "0/prop1/prop2"
    val = {"more": "props"}
    context = {"Pid": 1}
    root = {"participants": [context]}
    context_pointer = "/participants/0"


    First we see an `0` in pointer which denotes update
    should be within context. So no need to jump higher than context.

    So we truncate path to = "prop1/prop2"
    We see it's a string, so we know we are entering object's *prop1* as a property:
        {
            "participants": [{
                "prop1": ...
            }]
        }
    It's our whole Trial object, and ... here denotes our current descend.
    Now we truncate path one step further to = "prop2"
    Go down there and set `val={"more": "props"}` :
        {
            "participants": [{
                "prop1": {
                    "prop2": {"more": "props"}
                }
            }]
        }
    While context is a sub-part of that:
            {
                "prop1": {
                    "prop2": {"more": "props"}
                }
            }

    Args:
        pointer: relative jsonpointer to the property being set within current `context`
        val: the value being set
        context: the python object relatively to which val is being set
        root: the whole python object being constructed, contains `context`
        context_pointer: jsonpointer of `context` within `root`. Needed to jump up.
        verb: indicates if debug logic should be printed.
    Returns:
       Nothing
    """

    # don't do anything for None value, asserting None is permissable
    # is handled by the "allow_empty" functionality.
    if val is None:
        return

    # special case to set context doc itself
    if pointer.rstrip("#") == "":
        context.update(val)
        return

    # fill defaults
    if root is None:
        root = context
    if context_pointer is None:
        context_pointer = "/"

    # first we need to convert pointer to an absolute one
    # if it was a relative one (https://tools.ietf.org/id/draft-handrews-relative-json-pointer-00.html)
    if pointer.startswith("/"):
        jpoint = JsonPointer(pointer)
        doc = context

    else:
        # parse "relative" jumps up
        jumpups, slash, rem_pointer = pointer.partition("/")
        try:
            jumpups = int(jumpups.rstrip("#"))
        except ValueError:
            jumpups = 0

        # check that we don't have to jump up more than we dived in already
        assert jumpups <= context_pointer.rstrip("/").count(
            "/"
        ), f"Can't set value for pointer {pointer} to many jumps up from current context."

        # and we'll go down remaining part of `pointer` from there
        jpoint = JsonPointer(slash + rem_pointer)
        if jumpups > 0:
            # new context pointer
            higher_context_pointer = "/".join(
                context_pointer.strip("/").split("/")[: -1 * jumpups]
            )
            # making jumps up, by going down context_pointer but no all the way down
            if higher_context_pointer == "":
                doc = root
                assert (
                    len(jpoint.parts) > 0
                ), f"Can't update root object (pointer {pointer})"
            else:
                try:
                    doc = resolve_pointer(root, "/" + higher_context_pointer)
                except Exception as e:
                    raise Exception(e)
        else:
            doc = context

    # then we update it
    for i, part in enumerate(jpoint.parts[:-1]):

        try:
            doc = jpoint.walk(doc, part)

        except (JsonPointerException, IndexError) as e:
            # means that there isn't needed sub-object in place, so create one

            # look ahead to figure out a proper type that needs to be created
            next_thing = __jpointer_get_next_thing(jpoint.parts[i + 1])

            # insert it
            __jpointer_insert_next_thing(doc, jpoint, part, next_thing)

            # and `walk` it again - this time should be OK
            doc = jpoint.walk(doc, part)

        if isinstance(doc, EndOfList):
            actual_doc = __jpointer_get_next_thing(jpoint.parts[i + 1])
            __jpointer_insert_next_thing(doc.list_, jpoint, "-", actual_doc)
            doc = actual_doc

    __jpointer_insert_next_thing(doc, jpoint, jpoint.parts[-1], val)


def __jpointer_get_next_thing(next_part) -> Union[dict, list]:
    """
    Looking at next part of pointer creates a proper object - dict or list
    to insert into a doc, that is being `jsonpointer.walk`ed.
    """
    # `next_part` looks like array index like"[0]" or "-" (RFC 6901)
    if next_part == "-" or JsonPointer._RE_ARRAY_INDEX.match(str(next_part)):
        # so create array
        return []
    # or just dict as default
    else:
        return {}


def __jpointer_insert_next_thing(doc, jpoint, part, next_thing):
    """
    Puts next_thing into a doc (that is being `jsonpointer.walk`ed)
    by *part* "address". *jpoint* is Jsonpointer that is walked.
    """

    if part == "-":
        doc.append(next_thing)
    else:
        # part will return str or int, so we can use it for `doc[typed_part]`
        # and it will work for both - doc being `dict` or `array`
        typed_part = jpoint.get_part(doc, part)
        try:

            # check if something is there.
            if isinstance(typed_part, int):

                if isinstance(doc[typed_part], dict):

                    # merge the dictionaries.
                    doc[typed_part].update(next_thing)

                else:
                    # assign it
                    doc[typed_part] = next_thing

            else:
                # assign it
                doc[typed_part] = next_thing

        # if doc is an empty array we hit an error when we try to paste to [0] index,
        # so just append
        except IndexError:
            doc.append(next_thing)


def _get_recursively(search_dict, field):
    """
    Takes a dict with nested lists and dicts,
    and searches all dicts for a key of the field
    provided.
    """
    fields_found = []

    for key, value in search_dict.items():

        if key == field:
            fields_found.append(value)

        elif isinstance(value, dict):
            results = _get_recursively(value, field)
            for result in results:
                fields_found.append(result)

        elif isinstance(value, list):
            for item in value:
                if isinstance(item, dict):
                    more_results = _get_recursively(item, field)
                    for another_result in more_results:
                        fields_found.append(another_result)

    return fields_found


SUPPORTED_ASSAYS = [
    "wes_fastq",
    "wes_bam",
    "olink",
    "cytof",
    "ihc",
    "elisa",
    "rna_fastq",
    "rna_bam",
]

SUPPORTED_SHIPPING_MANIFESTS = [
    "pbmc",
    "plasma",
    "tissue_slide",
    "normal_blood_dna",
    "normal_tissue_dna",
    "tumor_tissue_dna",
    "h_and_e",
]
# weird non shipping manifest
SUPPORTED_WEIRD_MANIFESTS = ["tumor_normal_pairing"]
SUPPORTED_MANIFESTS = SUPPORTED_SHIPPING_MANIFESTS + SUPPORTED_WEIRD_MANIFESTS

SUPPORTED_ANALYSES = ["cytof_analysis", "wes_analysis"]

SUPPORTED_TEMPLATES = SUPPORTED_ASSAYS + SUPPORTED_MANIFESTS + SUPPORTED_ANALYSES

LocalFileUploadEntry = namedtuple(
    "LocalFileUploadEntry",
    ["local_path", "gs_key", "upload_placeholder", "metadata_availability"],
)


def _process_property(
    key: str,
    raw_val,
    key_lu: dict,
    data_obj: dict,
    format_context: dict,
    root_obj: Union[None, dict] = None,
    data_obj_pointer: Union[None, str] = None,
    verb: bool = False,
) -> List[LocalFileUploadEntry]:
    """
    Takes a single property (key, val) from spreadsheet, determines
    where it needs to go in the final object, then inserts it.
    Args:
        key: property name,
        raw_val: value of a property being processed,
        key_lu: dictionary to translate from template naming to json-schema
                property names
        data_obj: dictionary we are building to represent data
        format_context: dictionary of everything needed for 'gcs_uri_format'
                        in case this property has that
        root_obj: root dictionary we are building to represent data,
                  that holds 'data_obj' within 'data_obj_pointer'
        data_obj_pointer: pointer of 'data_obj' within 'root_obj'.
                          this will allow to process relative json-pointer properties
                          to jump out of data_object
        verb: boolean indicating verbosity
    Returns:
        [ LocalFileUploadEntry(
            local_path = "/local/file/from/excel/file/cell",
            gs_key = "constructed/GCS/path/where/this/artifact/should/endup",
            upload_placeholder = 'uuiduuiduuid-uuid-uuid-uuiduuid' # unique artifact/upload_placeholder,
            metadata_availability = boolean to indicate whether LocalFileUploadEntry should be extracted for metadata files
        ) ]
    """

    if verb:
        print(f"    processing property {key!r} - {raw_val!r}")
    # coerce value
    try:
        field_def = key_lu[key.lower()]
    except Exception:
        raise ParsingException(f"Unexpected property {key!r}.")

    if verb:
        print(f"      found def {field_def}")

    changes, files = _process_field_value(
        key=key,
        raw_val=raw_val,
        field_def=field_def,
        format_context=format_context,
        verb=verb,
    )

    for ch in changes:
        _set_val(ch.pointer, ch.value, data_obj, root_obj, data_obj_pointer, verb=verb)

    return files


class _AtomicChange(NamedTuple):
    """
    Represents exactly one "value set" operation on some data object
    `Pointer` being a json-pointer string showing where to set `value` to.
    """

    pointer: str
    value: Any


def _process_field_value(
    key: str, raw_val, field_def: dict, format_context: dict, verb: bool = False
) -> Tuple[List[_AtomicChange], List[LocalFileUploadEntry]]:
    """
    Processes one field value based on field_def taken from a ..._template.json schema.
    Calculates a list of `_AtomicChange`s within a context object
    and a list of file upload entries.
    A list of values and not just one value might arise from a `process_as` section
    in template schema, that allows for multi-processing of a single cell value.
    """

    # skip nullable
    if field_def.get("allow_empty"):
        if raw_val is None:
            return [], []

    # or set/update value in-place in data_obj dictionary
    pointer = field_def["merge_pointer"]
    if field_def.get("is_artifact") == 1:
        pointer += "/upload_placeholder"

    try:
        val, files = _calc_val_and_files(raw_val, field_def, format_context, verb=verb)
    except ParsingException:
        raise
    except Exception as e:
        raise ParsingException(
            f"Can't parse {key!r} value {str(raw_val)!r}: {e}"
        ) from e

    changes = [_AtomicChange(pointer, val)]

    # "process_as" allows to define additional places/ways to put that match
    # somewhere in the resulting doc, with additional processing.
    # E.g. we need to strip cimac_id='CM-TEST-0001-01' to 'CM-TEST-0001'
    # and put it in this sample parent's cimac_participant_id
    if "process_as" in field_def:
        for extra_fdef in field_def["process_as"]:
            # Calculating new "raw" val.
            extra_fdef_raw_val = raw_val
            # `eval` should be fine, as we're controlling the code argument in templates
            if "parse_through" in extra_fdef:
                try:
                    extra_fdef_raw_val = eval(extra_fdef["parse_through"])(raw_val)

                # catching everything, because of eval
                except:
                    extra_field_key = extra_fdef["merge_pointer"].rsplit("/", 1)[-1]
                    raise ParsingException(
                        f"Cannot extract {extra_field_key} from {key} value: {raw_val!r}"
                    )

            # recursive call
            extra_changes, extra_files = _process_field_value(
                key=key,
                raw_val=extra_fdef_raw_val,  # new "raw" val
                field_def=extra_fdef,  # merged field_def
                format_context=format_context,
                verb=verb,
            )

            files.extend(extra_files)
            changes.extend(extra_changes)

    return changes, files


def _get_file_ext(fname):
    return fname.rsplit(".")[-1]


def _format_single_artifact(
    local_path: str, uuid: str, field_def: dict, format_context: dict
):
    try:
        gcs_uri_format = field_def["gcs_uri_format"]
    except KeyError as e:
        raise KeyError(f"Empty gcs_uri_format for {field_def['key_name']!r}") from e

    if isinstance(gcs_uri_format, dict):
        if "check_errors" in gcs_uri_format:
            # `eval` should be fine, as we're controlling the code argument in templates
            err = eval(gcs_uri_format["check_errors"])(local_path)
            if err:
                raise ParsingException(err)

        try:
            gs_key = eval(gcs_uri_format["format"])(local_path, format_context)
        except Exception as e:
            raise ValueError(
                f"Can't format gcs uri for {field_def['key_name']!r}: {gcs_uri_format['format']}"
            )

    elif isinstance(gcs_uri_format, str):
        try:
            gs_key = gcs_uri_format.format_map(format_context)
        except KeyError as e:
            raise KeyError(
                f"Can't format gcs uri for {field_def['key_name']!r}: {gcs_uri_format}"
            )

        expected_extension = _get_file_ext(gs_key)
        provided_extension = _get_file_ext(local_path)
        if provided_extension != expected_extension:
            raise ParsingException(
                f"Expected {'.' + expected_extension} for {field_def['key_name']!r} but got {'.' + provided_extension!r} instead."
            )

    return LocalFileUploadEntry(
        local_path=local_path,
        gs_key=gs_key,
        upload_placeholder=uuid,
        metadata_availability=field_def.get("extra_metadata"),
    )


def _calc_val_and_files(raw_val, field_def: dict, format_context: dict, verb: bool):
    """
    Processes one field value based on field_def taken from a ..._template.json schema.
    Calculates a value and (if there's 'is_artifact') a file upload entry.
    """

    coerce = field_def["coerce"]
    val = coerce(raw_val)
    files = []

    if not field_def.get("is_artifact"):
        return val, files  # no files if it's not an artifact

    # deal with multi-artifact
    if field_def["is_artifact"] == "multi":
        if verb:
            print(f"      collecting multi local_file_path {field_def}")

        # In case of is_aritfact=multi we expect the value to be a comma-separated
        # list of local_file paths (that we will convert to uuids)
        # and also for the corresponding DM schema to be an array of artifacts
        # that we will fill with upload_placeholder uuids

        # So our value is a list of artifact placeholders
        val = []

        # and we iterate through local file paths:
        for num, local_path in enumerate(raw_val.split(",")):
            # Ignoring errors here as we're sure `coerce` will just return a uuid
            file_uuid = coerce(local_path)

            val.append({"upload_placeholder": file_uuid})

            files.append(
                _format_single_artifact(
                    local_path=local_path,
                    uuid=file_uuid,
                    field_def=field_def,
                    format_context=dict(
                        format_context,
                        num=num  # add num to be able to generate
                        # different gcs keys for each multi-artifact file.
                    ),
                )
            )

    else:
        if verb:
            print(f"      collecting local_file_path {field_def}")

        files.append(
            _format_single_artifact(
                local_path=raw_val,
                uuid=val,
                field_def=field_def,
                format_context=format_context,
            )
        )

    return val, files


class ParsingException(ValueError):
    pass


PRISM_PRISMIFY_STRATEGIES = {"overwriteAny": strategies.Overwrite()}


def prismify(
    xlsx: XlTemplateReader, template: Template, verb: bool = False
) -> (dict, List[LocalFileUploadEntry], List[Union[Exception, str]]):
    """
    Converts excel file to json object. It also identifies local files
    which need to uploaded to a google bucket and provides some logic
    to help build the bucket url.
    e.g. file list
    [
        {
            'local_path': '/path/to/fwd.fastq',
            'gs_key': '10021/CTTTPPPSS/wes_forward.fastq'
        }
    ]
    Args:
        xlsx: cidc_schemas.template_reader.XlTemplateReader instance
        template: cidc_schemas.template.Template instance
        verb: boolean indicating verbosity
    Returns:
        (tuple):
            arg1: clinical trial object with data parsed from spreadsheet
            arg2: list of LocalFileUploadEntries that describe each file identified:
                LocalFileUploadEntry(
                    local_path = "/local/path/to/a/data/file/parsed/from/template",
                    gs_key = "constructed/relative/to/clinical/trial/GCS/path",
                    upload_placeholder = "random_uuid-for-artifact-upload",
                    metadata_availability = boolean to indicate whether LocalFileUploadEntry should be extracted for metadata files
                )
            arg3: list of errors
    Process:
    * checks out `prism_preamble_object_pointer` which is a "standard"/absolute
    rfc6901 json-pointer from CT root object to a new assay location.
    E.g. for WES it is `/assays/wes/0`, in DeepDiff terms `ct["assays"]["wes"][0]`
    * creates such "parent/preamble" object.
    E.g. for WES an object that corresponds to a wes_assay will be created:
        {
          "assays": {
            "wes": [
              {
                ...    # we're here - this is "preamble" obj = "assay" obj
              }
            ]
          }
        }
    * then processes all "preamble_rows" properties from "..._template.json"
    to fill object's properties. It uses "merge_pointer"s relative to this
    "parent/preamble" object to determine exact location where to set value.
    In most cases it's just "0/field_name". Where "0" denotes that "field_name"
    is a field in the current object.
    With exceptions like - "3/protocol_identifier" which says basically
    "go 3 levels up in the hierarchy and take protocol_identifier field of the root".
    E.g. WES:
        {
          "protocol_identifier": "4412" # from `3/protocol_identifier`
          "assays": {
            "wes": [
              {
                "assay_creator": "DFCI" # from `0/assay_creator`
              }
            ]
          }
        }
    * then it goes in a loop over all "record" rows in .xlsx, and creates
    an object within that "parent" object for each row. These "record-objects"
    are created at "prism_data_object_pointer" location relative to "preamble".

    E.g. for WES: `"prism_data_object_pointer" : "/records/-"`
        {
          "assays": {
            "wes": [
              {
                "assay_creator": "DFCI",
                "records": [
                  {
                    ...    # we're here - this is "record" obj = "assay entry" obj
                  }
                ]
              }
            ]
          }
        }
    NB Minus sign at the end of "/records/-" is a special relative-json-pointer
    notation that means we need to create new object in an 'record' array.
    So it's like if python's `l.append(v)` would've been `l[-] = v`.
    * Prism now uses those "merge_pointer" relative to this "record" object,
    to populate field values of a "record" in the same way as with "preamble".
    E.g. for WES: `"prism_data_object_pointer" : "/records/-"`
        {
          "assays": {
            "wes": [
              {
                "assay_creator": "DFCI",
                "records": [
                  {
                    "cimac_id": ...                 # from "0/cimac_id",
                    "enrichment_vendor_lot": ...    # from "0/enrichment_vendor_lot",
                    "capture_date": ...             # from "0/capture_date",
                  }
                ]
              }
            ]
          }
        }
    * Finally, as there were many "records" object created/populated,
    Prism now uses `prism_preamble_object_schema` to merge all that together
    with respect to `mergeStrategy`es defined in that schema.
    """

    if template.type not in SUPPORTED_TEMPLATES:
        raise NotImplementedError(
            f"{template.type!r} is not supported, only {SUPPORTED_TEMPLATES} are."
        )

    errors_so_far = []

    # get the root CT schema
    root_ct_schema_name = (
        template.schema.get("prism_template_root_object_schema")
        or "clinical_trial.json"
    )
    root_ct_schema = load_and_validate_schema(root_ct_schema_name)
    # create the result CT dictionary
    root_ct_obj = (
        {f"__prism_origin__:///templates/{template.type}": "as root_ct_obj"}
        if verb
        else {}
    )
    template_root_obj_pointer = template.schema.get(
        "prism_template_root_object_pointer", ""
    )
    if template_root_obj_pointer != "":
        template_root_obj = (
            {f"__prism_origin__:///templates/{template.type}": "as template_root_obj"}
            if verb
            else {}
        )
        _set_val(template_root_obj_pointer, template_root_obj, root_ct_obj, verb=verb)
    else:
        template_root_obj = root_ct_obj

    # and merger for it
    root_ct_merger = Merger(root_ct_schema, strategies=PRISM_PRISMIFY_STRATEGIES)
    # and where to collect all local file refs
    collected_files = []

    # loop over spreadsheet worksheets
    for ws_name, ws in xlsx.grouped_rows.items():
        if verb:
            print(f"next worksheet {ws_name!r}")

        # Here we take only first two cells from preamble as key and value respectfully,
        # lowering keys to match template schema definitions.
        preamble_context = dict(
            (r.values[0].lower(), r.values[1]) for r in ws.get(RowType.PREAMBLE, [])
        )
        # We need this full "preamble dict" (all key-value pairs) prior to processing
        # properties from data_columns or preamble wrt template schema definitions, because
        # there can be a 'gcs_uri_format' that needs to have access to all values.

        templ_ws = template.schema["properties"]["worksheets"].get(ws_name)
        if not templ_ws:
            if ws_name in template.ignored_worksheets:
                continue

            errors_so_far.append(f"Unexpected worksheet {ws_name!r}.")
            continue

        preamble_object_schema = load_and_validate_schema(
            templ_ws.get("prism_preamble_object_schema", root_ct_schema_name)
        )
        preamble_merger = Merger(
            preamble_object_schema, strategies=PRISM_PRISMIFY_STRATEGIES
        )
        preamble_object_pointer = templ_ws.get("prism_preamble_object_pointer", "")
        data_object_pointer = templ_ws["prism_data_object_pointer"]

        # creating preamble obj
        preamble_obj = (
            {f"__prism_origin__:///templates/{template.type}/{ws_name}": "as preamble"}
            if verb
            else {}
        )

        # Processing data rows first
        data = ws[RowType.DATA]
        if data:
            # get the data
            headers = ws[RowType.HEADER][0]

            # for row in data:
            for row in data:

                if verb:
                    print(f"  next data row {row!r}")

                # creating data obj
                data_obj = (
                    {
                        f"__prism_origin__:///templates/{template.type}/{ws_name}/{row.row_num}": "as data_obj"
                    }
                    if verb
                    else {}
                )
                copy_of_preamble = (
                    {
                        f"__prism_origin__:///templates/{template.type}/{ws_name}/{row.row_num}": "as copy_of_preamble"
                    }
                    if verb
                    else {}
                )
                _set_val(
                    data_object_pointer,
                    data_obj,
                    copy_of_preamble,
                    template_root_obj,
                    preamble_object_pointer,
                    verb=verb,
                )

                # We create this "data record dict" (all key-value pairs) prior to processing
                # properties from data_columns wrt template schema definitions, because
                # there can be a 'gcs_uri_format' that needs to have access to all values.
                local_context = dict(
                    zip([h.lower() for h in headers.values], row.values)
                )

                # create dictionary per row
                for key, val in zip(headers.values, row.values):

                    try:
                        # get corr xsls schema type
                        new_files = _process_property(
                            key,
                            val,
                            key_lu=template.key_lu,
                            data_obj=data_obj,
                            format_context=dict(
                                local_context, **preamble_context
                            ),  # combine contexts
                            root_obj=copy_of_preamble,
                            data_obj_pointer=data_object_pointer,
                            verb=verb,
                        )
                        if new_files:
                            collected_files.extend(new_files)
                    except ParsingException as e:
                        errors_so_far.append(e)

                if verb:
                    print("  merging preambles")
                    print(f"   {preamble_obj}")
                    print(f"   {copy_of_preamble}")
                preamble_obj = preamble_merger.merge(preamble_obj, copy_of_preamble)
                if verb:
                    print(f"    merged - {preamble_obj}")

        # Now processing preamble rows
        if verb:
            print(f"  preamble for {ws_name!r}")
        for row in ws[RowType.PREAMBLE]:
            try:
                # process this property
                new_files = _process_property(
                    row.values[0],
                    row.values[1],
                    key_lu=template.key_lu,
                    data_obj=preamble_obj,
                    format_context=preamble_context,
                    root_obj=root_ct_obj,
                    data_obj_pointer=template_root_obj_pointer
                    + preamble_object_pointer,
                    verb=verb,
                )
                # TODO we might want to use copy+preamble_merger here too,
                # to for complex properties that require mergeStrategy

                if new_files:
                    collected_files.extend(new_files)
            except ParsingException as e:
                errors_so_far.append(e)

        # Now pushing it up / merging with the whole thing
        copy_of_template_root = (
            {
                f"__prism_origin__:///templates/{template.type}/{ws_name}": "as copy_of_template_root"
            }
            if verb
            else {}
        )
        _set_val(
            preamble_object_pointer, preamble_obj, copy_of_template_root, verb=verb
        )
        if verb:
            print("merging root objs")
            print(f" {template_root_obj}")
            print(f" {copy_of_template_root}")
        template_root_obj = root_ct_merger.merge(
            template_root_obj, copy_of_template_root
        )
        if verb:
            print(f"  merged - {template_root_obj}")

    if template_root_obj_pointer != "":
        _set_val(template_root_obj_pointer, template_root_obj, root_ct_obj, verb=verb)
    else:
        root_ct_obj = template_root_obj

    return root_ct_obj, collected_files, errors_so_far


def _set_data_format(ct: dict, artifact: dict):
    """
    Discover the correct data format for the given artifact.
    Args:
        ct: a clinical trial object with artifact inserted
        artifact: a reference to the artifact object inserted in `ct`.
        NOTE: in-place updates to artifact must trigger in-place updates to `ct`.
    """
    # This is invalid for all artifact types, and will
    # deliberately trigger a validation error below.
    artifact["data_format"] = "[NOT SET]"

    validator: jsonschema.Draft7Validator = load_and_validate_schema(
        "clinical_trial.json", return_validator=True
    )

    for error in validator.iter_errors(ct):
        if not isinstance(error, jsonschema.exceptions.ValidationError):
            continue

        if error.validator != "const":
            continue

        if error.path[-1] != "data_format":
            continue

        if error.instance != artifact["data_format"]:
            continue

        # Since data_format is specified as a constant in the schema,
        # the validator_value on this exception will be the desired data format.
        artifact["data_format"] = error.validator_value
        return

    # data format was not set!


def _get_uuid_info(ct: dict, artifact_uuid: str) -> (dict, dict):
    # Using uuids to find path in CT where corresponding artifact is located. Uuids are unique.
    uuid_field_path = get_path(ct, artifact_uuid)

    # As "uuid_field_path" contains path to a field with uuid,
    # we're looking for an artifact that contains it, not the "string" field itself
    # That's why we need skip_last=1, to get 1 "level" higher
    # from 'uuid_field_path' field to it's parent - existing_artifact obj
    return get_source(ct, uuid_field_path, skip_last=1)


def merge_artifact(
    ct: dict,
    artifact_uuid: str,
    object_url: str,
    assay_type: str,
    file_size_bytes: int,
    uploaded_timestamp: str,
    crc32c_hash: Optional[str] = None,
    md5_hash: Optional[str] = None,
) -> (dict, dict, dict):
    """
    create and merge an artifact into the metadata blob
    for a clinical trial. The merging process is automatically
    determined by inspecting the gs url path.
    Args:
        ct: clinical_trial object to be searched
        artifact_uuid: artifact identifier
        object_url: the gs url pointing to the object being added
        file_size_bytes: integer specifying the number of bytes in the file
        uploaded_timestamp: time stamp associated with this object
        md5_hash: md5 hash of the uploaded object, provided by GCS for non-composite objects
        crc32c_hash: crc32c hash of the uploaded object, provided by GCS for all objects
    Returns:
        ct: updated clinical trial object
        artifact: updated artifactf
        additional_artifact_metadata: relevant metadata collected while updating artifact
    """
    assert (
        crc32c_hash or md5_hash
    ), f"Either crc32c_hash or md5_hash must be provided for artifact: {object_url}"

    # urls are created like this in _process_property:
    file_name, uuid = object_url.split("/")[-2:]

    artifact_patch = {
        # TODO 1. this artifact_category should be filled out during prismify
        "artifact_category": "Assay Artifact from CIMAC",
        "object_url": object_url,
        "file_name": file_name,
        "file_size_bytes": file_size_bytes,
        "uploaded_timestamp": uploaded_timestamp,
    }

    if crc32c_hash:
        artifact_patch["crc32c_hash"] = crc32c_hash
    if md5_hash:
        artifact_patch["md5_hash"] = md5_hash

    return _update_artifact(ct, artifact_patch, artifact_uuid)


class InvalidMergeTargetException(ValueError):
    """Exception raised for target of merge_clinical_trial_metadata being non schema compliant."""


def merge_artifact_extra_metadata(
    ct: dict, artifact_uuid: str, assay_hint: str, extra_metadata_file: BinaryIO
) -> (dict, dict, dict):
    """
    Merges parsed extra metadata returned by extra_metadata_parsing to
    corresponding artifact objects within the patch.
    Args:
        ct: preliminary patch from upload_assay
        artifact_uuid: passed from upload assay
        assay_hint: assay type
        extra_metadata_file: extra metadata file in BinaryIO
    Returns:
        ct: updated clinical trial object
        artifact: updated artifact
        additional_artifact_metadata: relevant metadata collected while updating artifact
    """

    if assay_hint not in _EXTRA_METADATA_PARSERS:
        raise Exception(f"Assay {assay_hint} does not support extra metadata parsing")
    extract_metadata = _EXTRA_METADATA_PARSERS[assay_hint]

    artifact_extra_md_patch = extract_metadata(extra_metadata_file)
    return _update_artifact(ct, artifact_extra_md_patch, artifact_uuid)


def _wes_pipeline_config(
    upload_type: str,
) -> Callable[[dict, dict, str], Dict[str, str]]:
    """
        Returns a function that for specified assay type will
        generate a snakemake .yaml for each tumor/normal pair with
        .fastq or bam available. 

        upload_type == "assay" is for when the upload consists of .fastq/.bam assay data.
        So sample_id's for which we now have data should be in the `patch` arg,  
        and corresponding sample pairing info should be in the `full_ct` doc.

        Or if upload_type == "pairing" it's vice versa - new runs/pairs are calculated 
        from `patch` (which is a "pairing"), and then we look for sample data info in 
        the `full_ct`  

        In brush strokes it goes like this:
        - we need to figure out what new data we've got
        - based on that we calculate candidate analysis runs (tumor/normal pairs),
          that we might need to run now, due to new data arived.
        - last step actually check if for each candidate we have assay data
          for both - normal and tumor sample - and we render pipeline configs for those.

        As a result each assay or tumor_normal_pairing upload that "completes" some
        analysis run/pair (making all 3 pieces available - both samples data and a pairing)
        will be rendered only once - right after that upload.

        Patch is expected to be already merged into full_ct.
    """
    if upload_type not in ["assay", "pairing"]:
        raise NotImplementedError(
            f"Not supported type:{upload_type} for wes pipeline config generation."
        )

    curdir = os.path.dirname(os.path.abspath(__file__))
    loader = jinja2.FileSystemLoader(curdir)
    templ = jinja2.Environment(loader=loader).get_template(
        "pipeline_configs/wes_analysis_config.yaml.j2"
    )

    class AnalysisRun(NamedTuple):
        normal_cimac_id: str
        tumor_cimac_id: str
        run_id: str

    def internal(full_ct: dict, patch: dict, bucket: str) -> Dict[str, str]:
        """
            Generates a map from analysis run_ids found in full_ct
            to generated snakemake wes .yaml configs. 

            Patch is expected to be already merged into full_ct.
        """

        potential_new_runs: List[AnalysisRun] = []  # to be rendered

        if upload_type == "assay":
            # first we search for cimac_ids for which we just got new data files
            new_data_ids = set()
            # as we know that `patch` is a prism result of a wes upload
            # we are sure these getitem calls should be fine
            for wes in patch["assays"]["wes"]:
                for r in wes["records"]:
                    new_data_ids.add(r["cimac_id"])

            # then we compose a list of all the analysis runs
            # which are also cimac_ids tumor/normal pairs.
            for r in (
                full_ct.get("analysis", {})
                .get("wes_analysis", {})
                .get("pair_runs", [])
                # this notation is used due to that we don't know if we have any "analyses"
                # or any pairs in them. Thus we provide default empty containers.
            ):
                norm_i = r["normal"]["cimac_id"]
                tum_i = r["tumor"]["cimac_id"]
                runi = r["run_id"]
                # we filter runs, for which we have new data for at least on of the samples:
                if norm_i in new_data_ids or tum_i in new_data_ids:
                    potential_new_runs.append(AnalysisRun(norm_i, tum_i, runi))

        if upload_type == "pairing":
            # first we filter cimac_ids for which we now got pairing info
            # from analysis runs.
            potential_new_runs = [
                AnalysisRun(
                    r["normal"]["cimac_id"], r["tumor"]["cimac_id"], r["run_id"]
                )
                for r in patch["analysis"]["wes_analysis"]["pair_runs"]
            ]

        # then we compose a list of all records from all assay runs,
        # so we can filter out analysis runs for which we have data for both samples
        all_wes_records = {}
        for wes in full_ct["assays"]["wes"]:
            for r in wes["records"]:
                all_wes_records[r["cimac_id"]] = r

        res = {}
        for run in potential_new_runs:
            # if we have data files for *both* items in a tumor/normal pair
            if (
                run.normal_cimac_id in all_wes_records
                and run.tumor_cimac_id in all_wes_records
            ):
                # then this is a run we need, and so we render it
                res[run.run_id + ".yaml"] = templ.render(
                    **{
                        "run_id": run.run_id,
                        "tumor_sample": all_wes_records[run.normal_cimac_id],
                        "normal_sample": all_wes_records[run.tumor_cimac_id],
                        "BIOFX_BUCKET_NAME": bucket,
                    }
                )

        return res

    return internal


# This is a map from a assay type to a config generators,
# that should take (full_ct: dict, patch: dict, bucket: str) as arguments
# and return a map {"file_name": ["whatever pipeline config is"]}
_ANALYSIS_CONF_GENERATORS = {
    "wes_fastq": _wes_pipeline_config("assay"),
    "wes_bam": _wes_pipeline_config("assay"),
    "tumor_normal_pairing": _wes_pipeline_config("pairing"),
}


def generate_analysis_configs_from_upload_patch(
    ct: dict, patch: dict, template_type: str, bucket: str
) -> Dict[str, str]:
    """
    Generates all needed pipeline configs, from a new upload info.
    Args:
        ct: full metadata object *with `patch` merged*!
        patch: metadata patch passed from upload (assay or manifest)
        template_type: assay or manifest type
        bucket: a name of a bucket where data files are expected to be 
                available for the pipeline runner
    Returns:
        Filename to pipeline configs as a string map.
    """

    if template_type not in _ANALYSIS_CONF_GENERATORS:
        return {}

    return _ANALYSIS_CONF_GENERATORS[template_type](ct, patch, bucket)


def _update_artifact(
    ct: dict, artifact_patch: dict, artifact_uuid: str
) -> (dict, dict, dict):
    """ Updates the artifact with uuid `artifact_uuid` in `ct`,
    and return the updated clinical trial and artifact objects
    Args:
        ct: clinical trial object
        artifact_patch: artifact object patch
        artifact_uuid: artifact identifier
    Returns:
        ct: updated clinical trial object
        artifact: updated artifact
        additional_artifact_metadata: relevant metadata collected while updating artifact
    """

    artifact, additional_artifact_metadata = _get_uuid_info(ct, artifact_uuid)

    # TODO this might be better with merger:
    # artifact_schema = load_and_validate_schema(f"artifacts/{artifact_type}.json")
    # artifact_parent[file_name] = Merger(artifact_schema).merge(existing_artifact, artifact)
    artifact.update(artifact_patch)

    _set_data_format(ct, artifact)

    # return the artifact that was merged and the new object
    return ct, artifact, additional_artifact_metadata


class MergeCollisionException(ValueError):
    pass


class ThrowOnCollision(strategies.Strategy):
    """
    Similar to the jsonmerge's built in 'discard' strategy,
    but throws an error if the value already exists. This is
    used to prevent updates in `merge_clinical_trial_metadata`.
    """

    def merge(self, walk, base, head, schema, meta, **kwargs):
        if base.is_undef():
            return head
        if base.val != head.val:
            prop = base.ref.rsplit("/", 1)[-1]
            raise MergeCollisionException(
                f"Found conflicting values for {prop}: {base.val} (current) != {head.val} (incoming). "
                "Updates are not currently supported."
            )
        return base

    def get_schema(self, walk, schema, **kwargs):
        return schema


PRISM_MERGE_STRATEGIES = {
    # This overwrites the default jsonmerge merge strategy for literal values.
    "overwrite": ThrowOnCollision(),
    # Alias the builtin jsonmerge overwrite strategy
    "overwriteAny": strategies.Overwrite(),
}


def merge_clinical_trial_metadata(patch: dict, target: dict) -> (dict, List[str]):
    """
    Merges two clinical trial metadata objects together
    Args:
        patch: the metadata object to add
        target: the existing metadata object
    Returns:
        arg1: the merged metadata object
        arg2: list of validation errors
    """

    validator = load_and_validate_schema("clinical_trial.json", return_validator=True)

    # first we assert original object is valid
    try:
        validator.validate(target)
    except jsonschema.ValidationError as e:
        raise InvalidMergeTargetException(
            f"Merge target is invalid: {target}\n{e}"
        ) from e

    # assert the un-mutable fields are equal
    # these fields are required in the schema
    # so previous validation assert they exist
    if patch.get(PROTOCOL_ID_FIELD_NAME) != target.get(PROTOCOL_ID_FIELD_NAME):
        raise InvalidMergeTargetException(
            "Unable to merge trials with different " + PROTOCOL_ID_FIELD_NAME
        )

    # merge the two documents
    merger = Merger(validator.schema, strategies=PRISM_MERGE_STRATEGIES)
    merged = merger.merge(target, patch)

    return merged, list(validator.iter_errors(merged))


# Build a regex from the CIMAC ID pattern in the schema
cimac_id_regex = re.compile(
    load_and_validate_schema("sample.json")["properties"]["cimac_id"]["pattern"]
)


def parse_elisa(xlsx: BinaryIO) -> dict:
    """
    Parses the given ELISA grand serology results file to extract a list of sample IDs.
    If the file is not valid NPX but still xlsx the function will
    return a dict containing an empty list. Sample IDs not conforming to the CIMAC ID
    format will be skipped. The function will pass along any IO errors.
    Args:
        xlsx: an opened NPX file
    Returns:
        arg1: a dict of containing list of sample IDs and number of samples
    """

    # load the file
    if type(xlsx) == str:
        raise TypeError(f"parse_npx only accepts BinaryIO and not file paths")

    workbook = openpyxl.load_workbook(xlsx)

    # extract data to python
    ids = []
    worksheet = workbook[workbook.sheetnames[0]]

    for i, row in enumerate(worksheet.iter_rows()):

        if i == 0:
            assert "CIMAC ID" == row[0].value
            continue

        val = row[0].value

        if val:
            if cimac_id_regex.match(val):
                ids.append(val)

    sample_count = len(ids)

    samples = {"samples": ids, "number_of_samples": sample_count}

    return samples


def parse_npx(xlsx: BinaryIO) -> dict:
    """
    Parses the given NPX file from olink to extract a list of sample IDs.
    If the file is not valid NPX but still xlsx the function will
    return a dict containing an empty list. Sample IDs not conforming to the CIMAC ID
    format will be skipped. The function will pass along any IO errors.
    Args:
        xlsx: an opened NPX file
    Returns:
        arg1: a dict of containing list of sample IDs and number of samples
    """

    # load the file
    if type(xlsx) == str:
        raise TypeError(f"parse_npx only accepts BinaryIO and not file paths")

    workbook = openpyxl.load_workbook(xlsx)

    # extract data to python
    ids = []
    for worksheet_name in workbook.sheetnames:

        # simplify.
        worksheet = workbook[worksheet_name]
        seen_onlinkid = False
        for i, row in enumerate(worksheet.iter_rows()):

            # extract values from row
            vals = [col.value for col in row]

            first_cell = vals[0]

            # skip empty
            if len(vals) == 0 or first_cell is None:
                continue

            # check if we are starting ids
            if first_cell == "OlinkID":
                seen_onlinkid = True
                continue

            # check if we are done.
            if first_cell == "LOD":
                break

            # get the identifier
            if seen_onlinkid:
                # check that it is a CIMAC ID
                if cimac_id_regex.match(first_cell):
                    ids.append(first_cell)

    sample_count = len(ids)

    samples = {"samples": ids, "number_of_samples": sample_count}

    return samples


_EXTRA_METADATA_PARSERS = {"olink": parse_npx, "elisa": parse_elisa}
