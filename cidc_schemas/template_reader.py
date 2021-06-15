# -*- coding: utf-8 -*-

"""Defines the `XlTemplateReader` class for reading/validating templates from Excel templates."""

import logging
from itertools import dropwhile, zip_longest
from typing import Dict, List, Tuple, Union, BinaryIO, NamedTuple, Optional
from warnings import filterwarnings

filterwarnings(
    action="ignore",
    category=UserWarning,
    message="Data Validation extension is not supported",
    module="openpyxl",
)

import openpyxl

from .template import Template
from .template_writer import RowType, row_type_from_string
from .json_validation import validate_instance

logger = logging.getLogger("cidc_schemas.template_reader")


# A template row is any tuple whose first member is a RowType
class TemplateRow(NamedTuple):
    row_num: int
    row_type: RowType
    values: Union[List, Tuple]


# A mapping from RowType to a list of rows with that type
RowGroup = Dict[RowType, list]


class ValidationError(Exception):
    pass


class XlTemplateReader:
    """
    Reader and validator for Excel templates.
    """

    def __init__(self, template: Dict[str, List[TemplateRow]]):
        """
        Initialize a template reader.

        Arguments:
            template {Dict[str, List[TemplateRow]]} -- a mapping from worksheet names to worksheet rows
        """
        self.template = template

        # Mapping from worksheet names to rows grouped by type
        self.grouped_rows: Dict[str, RowGroup] = self._group_worksheet_rows()

        self.visited_fields = set()

    @staticmethod
    def _clean_value_row(values):
        rev_values = values[::-1]
        clean = list(dropwhile(lambda v: v is None, rev_values))
        clean = [v.strip() if isinstance(v, str) else v for v in clean]
        return clean[::-1]

    @staticmethod
    def from_excel(xlsx_path: Union[str, BinaryIO]):
        """
        Initialize an Excel template reader from an excel file.

        Arguments:
          xlsx_path {Union[str, BinaryIO]} -- path to the Excel file or the open file itself.

        Returns:
            arg1: XlTemplateReader or None if errors
            arg2: list of errors
        """

        # Load the Excel file
        workbook = openpyxl.load_workbook(xlsx_path)

        template = {}
        errors = []
        for worksheet_name in workbook.sheetnames:
            if worksheet_name in Template.ignored_worksheets:
                continue

            worksheet = workbook[worksheet_name]
            rows = []
            header_width = 0
            for row_num, row in enumerate(worksheet.iter_rows(), start=1):
                # Convert to string and extract type annotation
                typ, *values = [col.value for col in row]
                row_type = row_type_from_string(typ)

                # If entire row is empty, skip it (this happens at the bottom of the data table, e.g.)
                if not any(values):
                    continue

                # If no recognized row type is found but the row has data, throw an error
                if not row_type:
                    raise ValidationError(
                        f"No recognized row type found in row {worksheet_name}/{row_num}.\n"
                        f"Add #skip to column A to skip this row.\n"
                        f"Add #title to column A if this is a preamble row.\n"
                        f"Add #preamble to column A if this is a preamble row.\n"
                        f"Add #header to column A if this is a data header row.\n"
                        f"Add #data to column A if this is a data row."
                    )

                # Filter empty cells from the end of the row
                if row_type == RowType.HEADER:
                    values = XlTemplateReader._clean_value_row(values)
                    header_width = len(values)

                # Filter empty cells from the end of a data row
                elif row_type == RowType.DATA:
                    if not header_width:
                        errors.append(
                            f"Encountered data row (#{row_num} in worksheet {worksheet_name!r}) before header row"
                        )

                    values = XlTemplateReader._clean_value_row(values)
                    if len(values) > header_width:
                        errors.append(
                            f"Encountered data row (#{row_num} in worksheet {worksheet_name!r}) wider than header row"
                        )

                elif row_type == RowType.PREAMBLE:
                    values = XlTemplateReader._clean_value_row(values)

                    # preamble rows are [key, value], therefore must have exactly two entries
                    # if the value is optional and missing (ie None), will only get [key] after _clean_value_row, so add back None
                    if len(values) == 1:
                        values.append(None)
                    elif len(values) != 2:
                        errors.append(
                            f"Encountered preamble row in worksheet with width {len(values)} (expected 2)"
                        )

                # Reassemble parsed row and add to rows
                new_row = TemplateRow(row_num, row_type, values)
                rows.append(new_row)
            template[worksheet_name] = rows

        return XlTemplateReader(template), errors

    def _group_worksheet_rows(self) -> Dict[str, RowGroup]:
        """Map worksheet names to rows grouped by row type"""
        grouped_rows = {}
        for name, rows in self.template.items():
            grouped_rows[name] = self._group_rows(rows)
        return grouped_rows

    @staticmethod
    def _group_rows(rows) -> RowGroup:
        """Group rows in a worksheet by their type annotation"""
        # Initialize mapping from row types to lists of row content
        row_groups: Dict[RowType, List] = {row_type: [] for row_type in RowType}

        for row in rows:
            row_groups[row.row_type].append(row)

        return row_groups

    def _get_schema(
        self, key: str, schema: Dict[str, dict]
    ) -> (Optional[dict], Optional[Exception]):
        """Try to find a schemas for the given template key"""
        entity_name = Template._process_fieldname(key)
        if entity_name not in schema:
            return None, f"Found unexpected column {entity_name!r}"
        # Add a note saying this field was accessed
        self.visited_fields.add(entity_name)
        return schema[entity_name], None

    def _get_data_schemas(
        self, row_groups, data_schemas: Dict[str, dict], ignore_unexpected: bool
    ) -> (List[dict], List[str]):
        """Transform data table into a list of entity name + schema pairs"""
        header_row = row_groups[RowType.HEADER][0]
        data_rows = row_groups[RowType.DATA]

        errors = []

        # Ensure every data row has the right number of entries
        n_columns = len(header_row.values)
        for data_row in data_rows:
            n_entries = len(XlTemplateReader._clean_value_row(data_row.values))
            if n_entries > n_columns:
                errors.append(
                    f"Row {data_row.row_num} has {n_entries - n_columns} unexpected extra values."
                )

        schemas = []
        for header in header_row.values:
            if header:
                sch, err = self._get_schema(header, data_schemas)
                if err and not ignore_unexpected:
                    errors.append(err)
                schemas.append(sch)

        return schemas, errors

    def iter_errors(self, template: Template) -> List[str]:
        for name, schema in template.worksheets.items():
            yield from self._validate_worksheet(name, schema)

    def validate(self, template: Template) -> bool:
        """
        Validate a populated Excel template against a template schema.

        Arguments:
            template {Template} -- a template object containing the expected structure of the template
        Raises:
            ValidationError -- if the .xlsx file is invalid.
        Returns:
            True -- if everything is valid, otherwise raises an exception with validation reporting
        """

        invalid_messages = list(self.iter_errors(template))
        if invalid_messages:
            feedback = "\n".join(map(str, invalid_messages))
            raise ValidationError("\n" + feedback)

        return True

    def _make_validation_error(
        self, worksheet_name: str, field_name: str, row_num: int, message: str
    ) -> str:
        return f"Error in worksheet {worksheet_name!r}, row {row_num}, field {field_name!r}: {message}"

    def _validate_instance(self, value, schema):
        # All fields in an Excel template are required
        # except for "allow_empty"
        return validate_instance(
            value, schema, is_required=(not schema.get("allow_empty"))
        )

    def _validate_worksheet(self, worksheet_name: str, ws_schema: dict) -> List[str]:
        """Validate rows in a worksheet, returning a list of validation error messages."""
        self.visited_fields.clear()

        # If no worksheet is found, return only that error.
        if not worksheet_name in self.grouped_rows:
            yield f"Expected worksheet {worksheet_name!r} not found"
            return
        row_groups = self.grouped_rows[worksheet_name]

        if "preamble_rows" in ws_schema:
            # Validate preamble rows
            preamble_schemas = ws_schema["preamble_rows"]
            for row in row_groups[RowType.PREAMBLE]:
                key, value = row.values[0], row.values[1]
                if isinstance(key, str):
                    key = key.strip()
                if isinstance(value, str):
                    value = value.strip()

                schema, error = self._get_schema(key, preamble_schemas)
                if error:
                    yield self._make_validation_error(
                        worksheet_name, key, row.row_num, error
                    )
                    continue
                invalid_reason = self._validate_instance(value, schema)

                if invalid_reason:
                    yield self._make_validation_error(
                        worksheet_name, key, row.row_num, invalid_reason
                    )

            # Ensure that all preamble rows are present
            for name in preamble_schemas.keys():
                if name.strip() not in self.visited_fields:
                    yield (
                        f"Worksheet {worksheet_name!r} is missing expected template row: {name!r}"
                    )

        # Clear visited fields in case a data column has the same header as a preamble row
        self.visited_fields.clear()

        if "data_columns" in ws_schema:
            # Build up flat mapping of data schemas
            flat_data_schemas: Dict[str, dict] = {}
            for section in ws_schema["data_columns"].values():
                flat_data_schemas = {**flat_data_schemas, **section}

            # Validate data rows
            n_headers = len(row_groups[RowType.HEADER])
            if not n_headers == 1:
                yield (
                    f"Worksheet {worksheet_name!r}: Exactly one header row expected, but found {n_headers}"
                )
                return

            headers = row_groups[RowType.HEADER][0].values
            if not all(headers):
                yield (
                    f"Worksheet {worksheet_name!r}: Found an empty header cell at index {headers.index(None)}"
                )
                return
            headers = [h.strip() if isinstance(h, str) else h for h in headers]

            ignore_unexpected = bool(
                ws_schema.get("prism_arbitrary_data_merge_pointer")
            )
            data_schemas, errors = self._get_data_schemas(
                row_groups, flat_data_schemas, ignore_unexpected
            )
            for e in errors:
                yield (f"Worksheet {worksheet_name!r}: {e}")

            # Ensure that all data columns appear to be present
            for name in flat_data_schemas.keys():
                if name not in self.visited_fields:
                    yield (
                        f"Worksheet {worksheet_name!r} is missing expected template column: {name!r}"
                    )

            for data_row in row_groups[RowType.DATA]:
                for head, value, schema in zip_longest(
                    headers, data_row.values, data_schemas
                ):
                    if schema is None:
                        # we don't check unexpected column
                        # and we already have an error reported on that
                        continue
                    if isinstance(value, str):
                        value = value.strip()

                    invalid_reason = self._validate_instance(value, schema)

                    if invalid_reason:
                        yield self._make_validation_error(
                            worksheet_name, head, data_row.row_num, invalid_reason
                        )
