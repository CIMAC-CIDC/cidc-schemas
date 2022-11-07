"""Parsers for extracting extra metadata from files containing molecular data."""
import logging
import re
from codecs import BOM_UTF8
from typing import BinaryIO

import openpyxl
import pandas as pd

from ..json_validation import load_and_validate_schema

logger = logging.getLogger(__file__)

# Build a regex from the CIMAC ID pattern in the schema
cimac_id_regex = re.compile(
    load_and_validate_schema("sample.json")["properties"]["cimac_id"]["pattern"]
)
cimac_partid_regex = re.compile(
    load_and_validate_schema("participant.json")["properties"]["cimac_participant_id"][
        "pattern"
    ]
)


def parse_elisa(xlsx: BinaryIO) -> dict:
    """
    Parses the given ELISA grand serology results file to extract a list of sample IDs.
    If the file is not valid NPX but still xlsx the function will
    return a dict containing an empty list. Sample IDs not conforming to the CIMAC ID
    format will be skipped. The function will pass along any IO errors.
    Args:
        xlsx: an opened NPX file
    Returns:
        arg1: a dict of containing list of sample IDs and number of samples
    """

    # load the file
    if type(xlsx) == str:
        raise TypeError(f"parse_npx only accepts BinaryIO and not file paths")

    workbook = openpyxl.load_workbook(xlsx)

    # extract data to python
    ids = []
    worksheet = workbook[workbook.sheetnames[0]]

    idx = 0
    for i, row in enumerate(worksheet.iter_rows()):
        if i == 0:
            # find the one that looks like CIMAC ID
            # ignore case, switch underscores to spaces
            values = [
                str(i.value).upper().strip().replace("_", " ") if str(i.value) else ""
                for i in row
            ]
            assert any(["CIMAC ID" == i for i in values])
            idx = values.index("CIMAC ID")
            continue

        val = row[idx].value

        if val:
            if cimac_id_regex.match(val):
                ids.append(val)

    sample_count = len(ids)

    samples = {"samples": ids, "number_of_samples": sample_count}

    return samples


def parse_npx(xlsx: BinaryIO) -> dict:
    """
    Parses the given NPX file from olink to extract a list of sample IDs.
    If the file is not valid NPX but still xlsx the function will
    return a dict containing an empty list. Sample IDs not conforming to the CIMAC ID
    format will be skipped. The function will pass along any IO errors.
    Args:
        xlsx: an opened NPX file
    Returns:
        arg1: a dict of containing list of sample IDs and number of samples
    Raises:
        TypeError if xlsx is not a BinaryIO
        ValueError if the second row doesn't start with "NPX data"
    """

    # load the file
    if type(xlsx) == str:
        raise TypeError(f"parse_npx only accepts BinaryIO and not file paths")

    workbook = openpyxl.load_workbook(xlsx)

    # extract data to python
    ids = []
    for worksheet_name in workbook.sheetnames:

        # simplify.
        worksheet = workbook[worksheet_name]
        seen_onlinkid = False
        for i, row in enumerate(worksheet.iter_rows()):

            # extract values from row
            vals = [col.value for col in row]

            first_cell = vals[0]

            # skip empty
            if len(vals) == 0 or first_cell is None:
                continue

            # find OlinkID to locate the first data row
            if not seen_onlinkid:
                # check that this is actually an NPX file
                if i == 1 and first_cell != "NPX data":
                    raise ValueError("parse_npx got a file that is not in NPX format")

                # check if we are starting ids
                # use this to capture cases where the column name changes in spacing / capitalization
                ## needed because some data has 'OlinkID' while the standard seems to call for 'Olink ID'
                if str(first_cell).lower().replace(" ", "") == "olinkid":
                    seen_onlinkid = True
                    continue

            # once it's found keep getting ids until we're done
            else:
                # check if we are done.
                if first_cell == "LOD":
                    break

                # otherwise get the identifier
                # and check that it is a CIMAC ID
                if cimac_id_regex.match(first_cell):
                    ids.append(first_cell)

    sample_count = len(ids)

    samples = {"samples": ids, "number_of_samples": sample_count}

    return samples


def parse_clinical(file: BinaryIO) -> dict:
    """
    Parses the given clinical file to extract a list of participant IDs.
    By convention the first column should be "cimac_part_id" for files containing
    clinical data keyed to a specific participant. All tabs in each XLSX need to be checked
    however some tabs may contain supporting information so not having cimac_part_id is OK.
    Additionally some entire files may contain supporting information so not having any
    cimac_part_id is also OK.

    Also clinical data may contain information for particpants with no CIMAC IDs. For now
    these are simply skipped in our counting.

    Args:
        file: an opened clinical data file, either xlsx or csv
    Returns:
        arg1: a dict of containing list of participant IDs and number of participants
    Raises:
        TypeError if file is not a BinaryIO
    """
    # load the file
    if type(file) == str:
        raise TypeError(f"parse_clinical only accepts BinaryIO and not file paths")

    ids = set()

    try:
        workbook = openpyxl.load_workbook(file)
        assert len(workbook.sheetnames) > 0
    except:

        # seek back to the beginning of the file
        file.seek(0)

        # if it starts with a version, just skip it
        # via API, pandas still reads it even if we don't seek back
        # so instead pass as skiprows
        firstline = file.readline()
        # handle an edge case where the file starts with a Byte Order Mark
        if firstline.startswith(BOM_UTF8):
            firstline = firstline[len(BOM_UTF8) :]
        skiprows: int = int(
            firstline.startswith(b'"version",') or firstline.startswith(b"version,")
        )
        file.seek(0)

        try:
            csv = pd.read_csv(file, skiprows=skiprows)
        except Exception as e:
            logger.error("Error parsing clinical file: could not read as Excel or CSV")
            if hasattr(file, "name"):
                logger.error(f"filename: {file.name}")
            logger.error(str(e), exc_info=True)
            return {}
        else:
            if "cimac_part_id" in csv.columns:
                for possible_id in csv["cimac_part_id"].unique():
                    if cimac_partid_regex.match(str(possible_id)):
                        ids.add(possible_id)
            else:
                logger.error(
                    "Error parsing clinical CSV file: no cimac_part_id column found"
                )
                logger.error(f"Only found: {', '.join(list(csv.columns))}")

    else:
        # extract data to python
        for worksheet_name in workbook.sheetnames:

            # simplify.
            worksheet = workbook[worksheet_name]

            # iterate through all possible columns to find all cimac_part_id's
            # title must be in top 2 rows
            for column in worksheet.iter_cols(1, worksheet.max_column):
                # also check second row in case of version row
                # won't match the regex and title will be ignored
                possible_titles = (
                    {column[0].value}
                    if len(column) == 1
                    else {cell.value for cell in column[:2]}
                )
                if "cimac_part_id" in possible_titles:
                    for cell in column:
                        # some participant ID's might be blank for
                        # participants not in the system already (skip these for now)
                        if cell.value == "" or not cell.value:
                            continue

                        # get the identifier
                        # check that it is a CIMAC PART ID
                        if cimac_partid_regex.match(str(cell.value)):
                            ids.add(cell.value)

    part_count = len(ids)

    parts = {"participants": list(ids), "number_of_participants": part_count}

    return parts


EXTRA_METADATA_PARSERS = {
    "olink": parse_npx,
    "elisa": parse_elisa,
    "clinical_data": parse_clinical,
}
